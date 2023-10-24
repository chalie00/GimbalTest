import dataclasses
import os.path
import time

import openpyxl
import xlwings as xw

from dataclasses import dataclass
from openpyxl.styles import Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.dimensions import ColumnDimension
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string, get_column_letter

import Constants
import Constants as cons
import Testcase as tc

wb = openpyxl.Workbook()
sh = wb.active


# Verifier Information
@dataclass()
class VerifierInfo:
    name: str
    date: str
    location: str
    model: str


# Product Information Data From User Typing
product_info: [{'name': str, 'fw': str}] = []


# Check Whether The Report File Is Or Not
def whether_check_report():
    if os.path.isfile(cons.report_path):
        print('Report file is exist')
        wb = openpyxl.load_workbook(cons.report_path)
        sh = wb.active
    else:
        print('Report file is not exist')
        wb = openpyxl.Workbook()
        sh = wb.active


# Get The Verifier Information From User
def get_user_info():
    # Type The Verifier information
    print('Please type the verifier info\n')
    veri_name = input('verifier name?\n')
    veri_date = cons.current_time_bdYHMS
    veri_location = input('Verifier Location?\n')
    veri_model = input('Verify Application Name and Version?')
    user_info = VerifierInfo(name=veri_name, date=veri_date,
                             location=veri_location, model=veri_model)

    return user_info


# Get The Product Information
def get_product_info():
    # Type The Verifier information
    print('Please type the product info\n')
    modules = ['Model', 'IR', 'EO', 'OP1', 'OP2', 'OP3']
    for module in modules:
        name = input(f' If there are no additional fields to fill in, type None\n {module} Name?')
        fw = input(f'{module} FW?')
        if name.lower() == 'none':
            return
        elif fw.lower() == 'none':
            return
        else:
            item = {'name': name, 'fw': fw}
            product_info.append(item)


# Make A Table With Column Title
def make_table_with_title(titles: [str], title_pos: {'column': int, 'row': int},
                          align: {'ver': str, 'hor': str}, data_area: {'start': str, 'end': str}):
    whether_check_report()
    for i, title in enumerate(titles):
        title_col = sh.cell(column=title_pos['column'] + i, row=title_pos['row'], value=title)
        title_col.alignment = Alignment(vertical=align['ver'], horizontal=align['hor'])
    add_table = Table(displayName=f'{titles[0]}', ref=f'{data_area["start"]}:{data_area["end"]}')
    style = TableStyleInfo(name='TableStyleMedium9', showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    add_table.tableStyleInfo = style
    sh.add_table(add_table)


# Fill In Information Data in Excel
def fill_user_product_info(user):
    user_info_fields = [
        ['Verifier', f'{user.name}'],
        ['Date', f'{user.date}'],
        ['Location', f'{user.location}'],
        ['Application', f'{user.model}']
    ]
    # TODO (Com) - Apply the information in Excel with multiple table
    # Create A User Information Table
    make_table_with_title(cons.user_info_title, cons.user_info_title_pos,
                          {'ver': 'center', 'hor': 'center'}, cons.user_data_area)
    # Create A Product Information Table
    make_table_with_title(cons.product_info_title, cons.product_info_title_pos,
                          {'ver': 'center', 'hor': 'center'}, cons.product_data_area)
    # Fill In The User Field
    for row, info in enumerate(user_info_fields):
        sh.cell(column=cons.user_info_title_pos['column'],
                row=cons.user_info_title_pos['row'] + 1 + row, value=info[0])
        sh.cell(column=cons.user_info_title_pos['column'] + 1,
                row=cons.user_info_title_pos['row'] + 1 + row, value=info[1])

    # Fill In The Product Field
    for i, product in enumerate(product_info, start=cons.product_info_title_pos['row']):
        sh.cell(column=cons.product_info_title_pos['column'],
                row=i + 1, value=f'{product["name"]}')
        sh.cell(column=cons.product_info_title_pos['column'] + 1,
                row=i + 1, value=f'{product["fw"]}')

    # Alignment The Table
    for column_cells in sh.columns:
        length = max(len(str(cell.value)) * 1.3 for cell in column_cells)
        sh.column_dimensions[column_cells[0].column_letter].width = length

    wb.save(cons.report_path)


# Set The Test File Format
# Title
def set_report_format(date, model):
    sh.title = f'{model}'
    # Set The Report Title
    title = sh.merge_cells(start_column=cons.report_title_info['start_col'],
                           start_row=cons.report_title_info['start_row'],
                           end_column=cons.report_title_info['end_col'],
                           end_row=cons.report_title_info['end_row'])
    title_name = sh.cell(column=cons.report_title_info['start_col'],
                         row=cons.report_title_info['start_row'],
                         value=cons.report_title_info['title'])
    title_name.alignment = Alignment(horizontal=cons.report_title_info['align'],
                                     vertical=cons.report_title_info['align'])
    title_name.font = cons.report_title_font
    border_around_merged_cell(sh, cons.report_title_info['start_col'],
                              cons.report_title_info['start_row'],
                              cons.report_title_info['end_col'],
                              cons.report_title_info['end_row'],
                              cons.report_title_info['border'])
    # Set The Testcase Title
    testcase_titles = [cons.test_pic_pos, cons.test_code_pos, cons.test_dep_pos, cons.test_pre_pos,
                       cons.test_step_pos, cons.test_expect_pos, cons.test_result_pos, cons.test_time_pos]

    for title in testcase_titles:
        test_title = sh.cell(column=title['col'], row=title['row'], value=title['value'])
        test_title.alignment = Alignment(horizontal='center', vertical='center')
        test_title.font = cons.test_title_font

    # Test Step Cell Merge
    sh.merge_cells(start_column=cons.test_step_pos['col'], start_row=cons.test_step_pos['row'],
                   end_column=cons.test_step_pos['col'] + 1, end_row=cons.test_step_pos['row'])

    # Test Step Cell Stretch
    stretch_cells = ['F', 'G', 'H', 'I']
    for cell in stretch_cells:
        sh.column_dimensions[cell].width = 30

    # Expect Result Merge
    sh.merge_cells(start_column=cons.test_expect_pos['col'], start_row=cons.test_expect_pos['row'],
                   end_column=cons.test_expect_pos['col'] + 1, end_row=cons.test_expect_pos['row'])

    # Create Excel File By Time
    wb.save(cons.report_path)


# Fill In The Testcase Field
def fill_in_testcase_with_generate_code(testcase, code_str):
    whether_check_report()
    tc.generate_code(testcase, code_str)
    tc.fill_in_testcase(sh, tc.power_testcase, tc.tc_base_pos)
    wb.save(cons.report_path)


# Set Only A Specified Border Of Merged Cell
def border_merged_cell_with_part(sheet, start_column, start_row, end_column, end_row, style, part):
    if part == 'top':
        for i in range(start_column, end_column + 1):
            # Set The Top Border
            cell_top = sheet.cell(column=start_column, row=start_row)
            cell_top.border = Border(top=Side(border_style=f'{style}'))
            start_column += 1
    elif part == 'bottom':
        for i in range(start_column, end_column + 1):
            # Set The Bottom Border
            cell_bottom = sheet.cell(column=start_column, row=end_row)
            cell_bottom.border = Border(bottom=Side(border_style=f'{style}'))
            start_column += 1
    elif part == 'left':
        for i in range(start_row, end_row + 1):
            # Set The Left Border
            cell_left = sheet.cell(column=start_column, row=start_row)
            cell_left.border = Border(left=Side(border_style=f'{style}'))
            start_row += 1
    elif part == 'right':
        # Set The Right Border
        cell_right = sheet.cell(column=end_column, row=start_row)
        cell_right.border = Border(right=Side(border_style=f'{style}'))
        start_row += 1


# Set A Whole Border Of Merged Cell
def border_around_merged_cell(sheet, start_column, start_row, end_column, end_row, style):
    loop_count_fir = 0
    # Set The Top and Bottom Border
    for i in range(start_column, end_column + 1):
        loop_count_fir = i
        # Set The Top Border
        cell_top = sheet.cell(column=start_column, row=start_row)
        cell_top.border = Border(top=Side(border_style=f'{style}'))
        # Set The Bottom Border
        cell_bottom = sheet.cell(column=start_column, row=end_row)
        cell_bottom.border = Border(bottom=Side(border_style=f'{style}'))
        start_column += 1

    start_column = start_column - loop_count_fir + 1
    loop_count_sec = 0
    # Set The Left and Right Border
    for i in range(start_row, end_row + 1):
        loop_count_sec = i
        # Set The Left Border
        cell_left = sheet.cell(column=start_column, row=start_row)
        cell_left.border = Border(left=Side(border_style=f'{style}'))
        # Set The Right Border
        cell_right = sheet.cell(column=end_column, row=start_row)
        cell_right.border = Border(right=Side(border_style=f'{style}'))
        start_row += 1

    # Retry border Set The LeftTop and RightBottom
    # Because When Setting The Merge, Only The Final Setting Seem to be Applied
    start_row = start_row - loop_count_sec + 1
    sheet.cell(column=start_column, row=start_row).border = Border(
        left=Side(border_style=f'{style}'),
        top=Side(border_style=f'{style}'),
    )
    sheet.cell(column=end_column, row=end_row).border = Border(
        right=Side(border_style=f'{style}'),
        bottom=Side(border_style=f'{style}'),
    )
    sheet.cell(column=end_column, row=start_row).border = Border(
        right=Side(border_style=f'{style}'),
        top=Side(border_style=f'{style}'),
    )
    sheet.cell(column=start_column, row=end_row).border = Border(
        left=Side(border_style=f'{style}'),
        bottom=Side(border_style=f'{style}'),
    )


# Return moved cell coordinate with coordinate
def moved_cell_coordinate_with_coordinate(coord, move_column, move_row):
    xy = coordinate_from_string(coord)
    x = column_index_from_string(xy[0]) + move_column
    y = xy[1] + move_row
    conx = get_column_letter(x)
    convert_coord = conx + f'{y}'
    return convert_coord


# Return moved tuple data with coordinate
def moved_cell_tuple_with_coordinate(coord, move_column, move_row):
    xy = coordinate_from_string(coord)
    x = column_index_from_string(xy[0]) + move_column
    y = xy[1] + move_row
    tuple_data = {'column': x, 'row': y}
    return tuple_data
