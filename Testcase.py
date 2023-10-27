import os

from openpyxl.drawing.spreadsheet_drawing import AnchorMarker

import Constants as cons
import ReportFormat as rf
import Major_Function as mf

import openpyxl

from openpyxl.styles import Alignment


tc_base_pos = {'col': 3, 'row': 21}
power_testcase = [
    {'depth': 'test -> test1 -> test2',
     'pre': 'test 상태 유지',
     'step': 'test에서 test1로 이동 후 test2를 선택한다.',
     'expect': '정상적으로 test2가 선택된다.',
     'code': '',
     'image': True
     },
    {'depth': 'test3 -> test4 -> test5',
     'pre': 'test3 상태 유지',
     'step': 'test3에서 test4로 이동 후 test5를 선택한다.',
     'expect': '정상적으로 test5가 선택된다.',
     'code': '',
     'image': True
     },
    {'depth': 'test8 -> test9 -> test10',
     'pre': 'test7 상태 유지',
     'step': 'test8에서 test9로 이동 후 test10를 선택한다.',
     'expect': '정상적으로 test10가 선택된다.',
     'code': '',
     'image': False
     },
]


# TODO (future) Generate depth

# TODO (future) Create a sheet for each Test Category


# TODO (Com 10/19) - Generate Test Code
def generate_code(testcase, code_str):
    case_count = int(len(testcase))
    for i in range(0, case_count):
        test_code = code_str + '-' + f'{i + 1:03d}'
        power_testcase[i]['code'] = test_code


# TODO (Com 10/20) - Fill In The Testcase for each Row
# TODO (Com 10/24) - Fill In Test Step Complete Time
def fill_in_testcase(sh, testcases, code_pos):
    for testcase in testcases:
        code = sh.cell(column=tc_base_pos['col'], row=tc_base_pos['row'],
                       value=f'{testcase["code"]}')
        depth = sh.cell(column=tc_base_pos['col'] + 1, row=tc_base_pos['row'],
                        value=f'{testcase["depth"]}')
        pre_state = sh.cell(column=tc_base_pos['col'] + 2, row=tc_base_pos['row'],
                            value=f'{testcase["pre"]}')
        step = sh.cell(column=tc_base_pos['col'] + 3, row=tc_base_pos['row'],
                       value=f'{testcase["step"]}')
        expect = sh.cell(column=tc_base_pos['col'] + 5, row=tc_base_pos['row'],
                         value=f'{testcase["expect"]}')
        current_time = sh.cell(column=tc_base_pos['col'] + 8, row=tc_base_pos['row'],
                               value=f'{cons.current_only_time}')
        tc_base_pos['row'] += 1
        aligns = [code, depth, pre_state, step, expect, current_time]
        for align in aligns:
            align.alignment = Alignment(horizontal='left', vertical='center')
    # Alignment The Table
    for column_cells in sh.columns:
        length = max(len(str(cell.value)) * 1.3 for cell in column_cells)
        sh.column_dimensions[column_cells[0].column_letter].width = length
        # TODO (future) Applying fonts in the future


# TODO (future) If image field is true, Attach a test result image
# TODO (future)  Applying in the future
def attach_image(testcases, path):
    fwd_cap_code:{'cap': bool, 'code': str} = []
    for i, testcase in enumerate(testcases):
        if testcase['image']:
            temp = {'cap': True, 'code': testcase['code']}
            fwd_cap_code.append(temp)


# TODO (Com 10/23) - Attach to report after capture
def attach_to_report_after_capture_img(dia, path, group, code, image_size_rate):
    # Image capture
    image_path = mf.set_save_image_dir(group, code)
    captured_img = dia.capture_as_image().save(image_path)
    img = openpyxl.drawing.image.Image(image_path)
    # Resize the image
    img.width = img.width * image_size_rate
    img.height = img.height * image_size_rate
    # Check the code num
    # Check the Excel coordinate
    if os.path.isfile(path):
        print('Report file is exist')
        wb = openpyxl.load_workbook(path)
        sh = wb.active
    else:
        print('Report file is not exist')
        wb = openpyxl.Workbook()
        sh = wb.active

    for row in sh.iter_rows():
        for cell in row:
            if cell.value == f'{code}':
                # adjust the image position in Excel sheet
                # attach the image to report
                pos = rf.moved_cell_coordinate_with_coordinate(cell.coordinate, -1, 0)
                sh.add_image(img, pos)
                # TODO (Com 10/24) - Adjust Cell Size to be the same size as the image
                cell_tuple_coord = rf.moved_cell_tuple_with_coordinate(cell.coordinate,
                                                               0, 0)
                cell_move_column = rf.moved_cell_coordinate_with_coordinate(cell.coordinate, -1, 0)
                sh.row_dimensions[cell_tuple_coord['row']].height = img.height - 15
                sh.column_dimensions[cell_move_column[0]].width = img.width - 140
    wb.save(path)


