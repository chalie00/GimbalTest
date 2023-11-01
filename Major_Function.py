import time

import numpy as np
import os.path

import pyautogui
import cv2
import openpyxl
import imgsim
import torchvision
import matplotlib.pyplot as plt

import Constants as cons

from pywinauto import WindowSpecification
from PIL import Image, ImageGrab
from pynput.mouse import Listener, Button
from mss import mss


def login_correct_WithOnlyPW(dia, pw):
    # dia['passwordEdit'].type_keys(f'{pw}')
    # dia['OK'].click()
    dia.passwordEdit.type_keys(pw)
    time.sleep(1)
    dia['OK'].click()


# Upload To The Excel File With Menus(can select object) Array
def upload_excel_with_menus(dia, menus):
    if os.path.isfile(cons.report_path):
        print('Report file is exist')
        wb = openpyxl.load_workbook(cons.report_path)
        sh = wb.active
        capture_after_check_menus_count(wb, sh, dia, menus)
    else:
        print('Report file is not exist')
        wb = openpyxl.Workbook()
        sh = wb.active
        capture_after_check_menus_count(wb, sh, dia, menus)


# Upload To The Excel File With Coordinate(can't select object) Array
# After Check The Excel File
def upload_excel_after_check_Report(dia: WindowSpecification, coordinates: [{str: str, str: [int]}]):
    if os.path.isfile(cons.report_path):
        wb = openpyxl.load_workbook(cons.report_path)
        sh = wb.active
        capture_upload_excel_with_coordinate(dia, wb, sh, coordinates)
    else:
        wb = openpyxl.Workbook()
        sh = wb.active
        capture_upload_excel_with_coordinate(dia, wb, sh, coordinates)


# Upload To The Excel File With Coordinate(can't select object) Array
def capture_upload_excel_with_coordinate(dia: WindowSpecification, wb: None, sh: None,
                                         coordinates: [{str: str, str: [int]}]):
    for coordinate in coordinates:
        area_capture_with_coordinate(dia, coordinate['name'], coordinate['coord'])
        sh.cell(row=cons.excel_No, column=1, value=f'{coordinate["name"]} Crop')
        img = openpyxl.drawing.image.Image(rf'Capture\{coordinate["name"]}_crop.png')
        sh.add_image(img, f'A{cons.excel_No + 1}')
        time.sleep(1)
        cons.excel_No += 3
    wb.save(cons.report_path)


# Check Menu Array Count And Capture
def capture_after_check_menus_count(wb, sh, dia, menus):
    if len(menus) > 1:
        for i, menu in enumerate(menus):
            dia.child_window(title=f'{menu}', control_type='CheckBox').click_input()
            dia.capture_as_image().save(f'Capture/{menu}.png')
            sh.cell(row=cons.excel_No, column=1, value=rf'{menus[i]}')
            img = openpyxl.drawing.image.Image(rf'Capture\{menus[i]}.png')
            sh.add_image(img, f'A{cons.excel_No + 1}')
            time.sleep(1)
            cons.excel_No += 35
        wb.save(cons.report_path)
    else:
        dia.capture_as_image().save(rf'Capture\{menus[0]}.png')
        time.sleep(1)
        sh.cell(row=cons.excel_No, column=1, value=f'{menus[0]}')
        img = openpyxl.drawing.image.Image(rf'Capture\{menus[0]}.png')
        sh.add_image(img, f'A{cons.excel_No + 1}')
        time.sleep(1)
        cons.excel_No += 35
        wb.save(cons.report_path)


#  Calculate coordinates from the zero position in the main frame
#  to the selected element
def cal_element_coordinate(left, top, right, bottom):
    ele_coord = {'l': 0,
                 't': 0,
                 'r': 0,
                 'b': 0}
    coordinate_array = cons.Dig_zero_Coordinate
    ele_coord['l'] = left - coordinate_array[0]
    ele_coord['t'] = top - coordinate_array[1]
    ele_coord['r'] = right - coordinate_array[0]
    ele_coord['b'] = bottom - coordinate_array[1]

    return ele_coord


# Capture selected Area or Element
def area_capture_with_coordinate(dia: WindowSpecification, title: str, coordinate: [int]):
    coord = cal_element_coordinate(coordinate[0], coordinate[1],
                                   coordinate[2], coordinate[3])
    img = dia.capture_as_image()
    (img.crop([coord['l'], coord['t'], coord['r'], coord['b']])
     .save(fr'Capture\{title}_crop.png'))


# Set the save directory of capture image
def set_save_image_dir(group, code):
    current_time = cons.current_time_bdYHMS
    save_path = rf'Capture\{group}\{code}_{current_time}.png'
    return save_path


# TODO (Com 10/24) - Virtual JoyStick Control By Mouse Drag
def mouse_drag_drop(start_coord, direction, hold_time, degree, end_coord):
    pyautogui.moveTo(start_coord[0], start_coord[1])
    pyautogui.mouseDown()
    cal_degree = [(abs(end_coord[0] - start_coord[0])) * degree,
                  (abs(start_coord[1] - end_coord[1])) * degree]
    print(cal_degree)
    if direction == 'up':
        pyautogui.moveTo(start_coord[0], start_coord[1] - cal_degree[1])
    elif direction == 'down':
        pyautogui.moveTo(start_coord[0], start_coord[1] + cal_degree[1])
    elif direction == 'left':
        pyautogui.moveTo(start_coord[0] - cal_degree[0], start_coord[1])
    elif direction == 'right':
        pyautogui.moveTo(start_coord[0] + cal_degree[0], start_coord[1])
    time.sleep(hold_time)
    pyautogui.mouseUp()


# TODO (com 10/27) - Record specify screen while running python code
# Capture the specify area of screen
def capture_video_with_area(sel_monitor: int, top: int, left: int, width: int, height: int,
                            rec_time: int, out_name: str, v_codec: str, fps: int):
    # Get the resolution of using user monitor
    screen_size = tuple(pyautogui.size())

    # Set the video codec
    set_codec = cv2.VideoWriter_fourcc(*f'{v_codec}')

    # Create a videoWriter Object
    out = cv2.VideoWriter(f'{out_name}.avi', set_codec, fps, (width, height))

    for i in range(int(rec_time * fps)):
        print('start recording')
        sct = mss()
        # Selecting a monitor to capture from multiple monitors
        mon = sct.monitors[sel_monitor]
        monitor = {'top': mon['top'] + top, 'left': mon['left'] + left,
                   'width': 1280, 'height': 750, 'mon': mon}
        # img = pyautogui.screenshot(region=(320, 165, 1280, 720))
        img = sct.grab(monitor)
        frame = np.array(img)
        frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
        out.write(frame)
        # cv2.imshow('screenshot', frame)
        if cv2.waitKey(1) == ord('q'):
            break
    print('end recording')
    out.release()
    cv2.destroyAllWindows()


# TODO (Com 10/30) - Compare a Image
def compare_img(base_img_path, target_img_path):
    vtr = imgsim.Vectorizer()
    base = cv2.imread(f'{base_img_path}')
    target = cv2.imread(f'{target_img_path}')

    base_vec = vtr.vectorize(base)
    target_vec = vtr.vectorize(target)
    compare_image = imgsim.distance(base_vec, target_vec)
    print(f'result = round({compare_image})')

    return round(compare_image, 2)


# TODO (Com 10/31) - Detect different part of image
def detect_image_diff(base_img_path, target_img_path):
    # read the image
    base = cv2.imread(base_img_path)
    target = cv2.imread(target_img_path)

    # Calculate a different part of image
    cal_diff = cv2.absdiff(base, target)

    # Convert GrayScale
    gray_trg = cv2.cvtColor(target, cv2.COLOR_BGR2GRAY)
    gray_cal_diff = cv2.cvtColor(cal_diff, cv2.COLOR_BGR2GRAY)

    # Convert to RGB from BGR
    base_rgb = cv2.cvtColor(base, cv2.COLOR_BGR2RGB)
    target_rgb = cv2.cvtColor(target, cv2.COLOR_BGR2RGB)

    # Normalization for applying color map
    norm_diff = gray_cal_diff / np.max(gray_cal_diff)

    # Weight the difference image to reflect the color of the second image
    diff_img = cv2.addWeighted(gray_trg, 0.1, gray_cal_diff, 2, 100)
    diff_colored = np.zeros_like(target_rgb)
    diff_colored[..., 0] = target_rgb[..., 0] * norm_diff
    diff_colored[..., 1] = target_rgb[..., 1] * norm_diff
    diff_colored[..., 2] = target_rgb[..., 2] * norm_diff

    image_info_arrays = [
        ['base image', 'off'],
        ['target image', 'off'],
        ['Difference(Grayscale)', 'off'],
        ['Difference(Colored)', 'off']
    ]
    image_arrays = [base_rgb, target_rgb, diff_img, diff_colored]

    # Display the image in plt
    align_plot_with_save(2, 2, (30, 30), image_info_arrays, image_arrays)


# TODO (Com 10/31) - Display a rectangle after detect different parts of image
def display_rect_after_detect(base_img_path, target_img_path):
    # read the image
    base = cv2.imread(base_img_path)
    target = cv2.imread(target_img_path)

    # Calculate a different part of image
    cal_diff = cv2.absdiff(base, target)

    # Convert GrayScale
    gray_cal_diff = cv2.cvtColor(cal_diff, cv2.COLOR_BGR2GRAY)

    # Convert to RGB from BGR
    base_rgb = cv2.cvtColor(base, cv2.COLOR_BGR2RGB)
    target_rgb = cv2.cvtColor(target, cv2.COLOR_BGR2RGB)

    # Normalization for applying color map
    norm_diff = gray_cal_diff / np.max(gray_cal_diff)

    target_diff = target_rgb.copy()
    contours, _ = cv2.findContours(gray_cal_diff, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    for contour in contours:
        x, y, w, h = cv2.boundingRect(contour)
        cv2.rectangle(target_diff, (x, y), (x + w, y + h), (255, 255, 0), 2)

    image_info_arrays = [
        ['base image', 'off'],
        ['target image', 'off'],
        ['Difference(Grayscale)', 'off'],
        ['Difference(Colored)', 'off']
    ]
    image_arrays = [base_rgb, target_rgb, gray_cal_diff, target_diff]

    # Display the image in plt
    align_plot_with_save(2, 2, (30, 30), image_info_arrays, image_arrays)


# TODO (Com 10/31) - Align a plot by axes, save fig image,
# TODO (Com 10/31) - Save an image with different parts of it marked as squares than another image
def align_plot_with_save(column: int, row: int, fig_size: (int, int), image_info: [[]], image_array):
    # Display the result by Matplotlib
    fig, axes = plt.subplots(row, column, figsize=fig_size)

    # Generate the number of cases with rows and columns
    axes_arrays: [[int, int]] = []
    for i in range(0, row):
        for k in range(0, column):
            num_cases = [i, k]
            axes_arrays.append(num_cases)

    # align the plot
    for i, axes_array in enumerate(axes_arrays, start=0):
        axes[axes_array[0], axes_array[1]].imshow(image_array[i])
        axes[axes_array[0], axes_array[1]].set_title(image_info[i][0])
        axes[axes_array[0], axes_array[1]].axis(image_info[i][1])

    plt.tight_layout()
    plt.show()

    # Save the image
    fig.savefig('Compare.png')
    bbox = axes[1, 1].get_tightbbox(fig.canvas.get_renderer())
    fig.savefig('Compare2.png', bbox_inches=bbox.transformed(fig.dpi_scale_trans.inverted()))


# Function called on a mouse click
# def on_click(x, y, button, pressed):
#     # Check if the left button was pressed
#     if pressed and button == Button.left:
#         # Print the click coordinates
#         print(f'x={x} and y={y}')
#
#
# # Initialize the Listener to monitor mouse clicks
# with Listener(on_click=on_click) as listener:
#     listener.join()
